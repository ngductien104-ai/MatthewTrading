"""Privacy-preserving reader for the local encrypted customer Excel exports."""

from __future__ import annotations

import hashlib
import json
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class PortfolioLookupError(ValueError):
    """Raised when an encrypted customer/account reference cannot be resolved."""


_FILE_KINDS = {
    "portfolio": "Portfolio_total_",
    "nav": "NAV_endperiod_total_",
    "daily": "Daily_report_total_",
    "toi": "TOI_total_",
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip().replace(" ", "").replace(",", "")
    try:
        return float(raw)
    except ValueError:
        return 0.0


def _norm_header(value: Any) -> str:
    return " ".join(_text(value).replace("\n", " ").split()).casefold()


def _date_from_filename(path: Path) -> datetime:
    matches = re.findall(r"(?<!\d)(\d{8})(?!\d)", path.name)
    for token in reversed(matches):
        try:
            return datetime.strptime(token, "%d%m%Y")
        except ValueError:
            continue
    return datetime.fromtimestamp(path.stat().st_mtime)


def _latest_files(root: Path) -> dict[str, Path]:
    if not root.is_dir():
        raise PortfolioLookupError("Customer Excel database directory is unavailable")
    files = [p for p in root.rglob("*") if p.is_file() and not p.name.startswith("~$")]
    selected: dict[str, Path] = {}
    for kind, prefix in _FILE_KINDS.items():
        candidates = [p for p in files if p.name.startswith(prefix)]
        if candidates:
            selected[kind] = max(candidates, key=lambda p: (_date_from_filename(p), p.stat().st_mtime))
    missing = sorted(set(_FILE_KINDS) - set(selected))
    if missing:
        raise PortfolioLookupError("Missing required customer export types: " + ", ".join(missing))
    return selected


def _read_rows(path: Path) -> list[dict[str, Any]]:
    # Some source exports contain XLSX bytes under a legacy .xls suffix.
    workbook = load_workbook(BytesIO(path.read_bytes()), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        raw_rows = sheet.iter_rows(values_only=True)
        headers: list[str] | None = None
        result: list[dict[str, Any]] = []
        for row in raw_rows:
            values = list(row)
            if headers is None:
                normalized = [_norm_header(v) for v in values]
                if "tài khoản" in normalized and "tên khách hàng" in normalized:
                    headers = normalized
                continue
            record = {header: values[i] if i < len(values) else None for i, header in enumerate(headers) if header}
            if _text(record.get("tài khoản")):
                result.append(record)
        if headers is None:
            raise PortfolioLookupError(f"Unrecognized workbook schema: {path.name}")
        return result
    finally:
        workbook.close()


def _matching_rows(
    rows: list[dict[str, Any]], account_token: str, customer_token: str | None
) -> list[dict[str, Any]]:
    account = account_token.strip()
    customer = customer_token.strip() if customer_token else None
    return [
        row
        for row in rows
        if _text(row.get("tài khoản")) == account
        and (customer is None or _text(row.get("tên khách hàng")) == customer)
    ]


def _pick(row: dict[str, Any] | None, name: str) -> float:
    return _number((row or {}).get(name))


class ExcelPortfolioRepository:
    """Resolve encrypted references to a de-identified portfolio snapshot."""

    def __init__(self, root: Path) -> None:
        self.root = root.resolve()

    def lookup(self, *, account_token: str, customer_token: str | None = None) -> dict[str, Any]:
        if not account_token.strip():
            raise PortfolioLookupError("Encrypted account lookup token is required")

        files = _latest_files(self.root)
        tables = {kind: _read_rows(path) for kind, path in files.items()}
        matches = {
            kind: _matching_rows(rows, account_token, customer_token)
            for kind, rows in tables.items()
        }
        if not matches["portfolio"] and not matches["nav"]:
            raise PortfolioLookupError("No portfolio was found for the supplied encrypted references")

        nav_row = matches["nav"][0] if matches["nav"] else None
        daily_row = matches["daily"][0] if matches["daily"] else None
        toi_row = matches["toi"][0] if matches["toi"] else None
        nav = _pick(nav_row, "nav cuối kỳ") or _pick(daily_row, "nav cuối kỳ")

        positions: list[dict[str, Any]] = []
        for row in matches["portfolio"]:
            market_value = _pick(row, "giá trị thị trường")
            positions.append({
                "symbol": _text(row.get("mã chứng khoán")),
                "tradable_quantity": _pick(row, "chứng khoán giao dịch"),
                "t1_quantity": _pick(row, "chứng khoán t1"),
                "t2_quantity": _pick(row, "chứng khoán t2"),
                "restricted_quantity": _pick(row, "chứng khoán phong tỏa"),
                "pledged_quantity": _pick(row, "chứng khoán thế chấp"),
                "market_price": _pick(row, "giá thị trường"),
                "market_value": market_value,
                "cost_price": _pick(row, "giá vốn"),
                "cost_value": _pick(row, "giá trị vốn"),
                "unrealized_pnl": _pick(row, "lãi lỗ"),
                "unrealized_pnl_pct": _pick(row, "% lãi lỗ"),
                "weight_of_nav": market_value / nav if nav else None,
            })

        account_fingerprint = hashlib.sha256(account_token.encode("utf-8")).hexdigest()[:12]
        as_of = max(_date_from_filename(path) for path in files.values()).date().isoformat()
        snapshot = {
            "portfolio_ref": f"portfolio-{account_fingerprint}",
            "as_of": as_of,
            "currency": "VND",
            "nav": nav,
            "cash": _pick(nav_row, "tổng tiền mặt cơ sở cuối kỳ"),
            "interest_support_cash": _pick(nav_row, "tổng tiền htls cuối kỳ"),
            "other_cash": _pick(nav_row, "tổng các loại tiền cuối kỳ khác"),
            "equity_market_value": _pick(nav_row, "giá trị chứng khoán cơ sở cuối kỳ"),
            "bond_market_value": _pick(nav_row, "giá trị trái phiếu cuối kỳ"),
            "fund_market_value": _pick(nav_row, "giá trị chứng chỉ quỹ cuối kỳ"),
            "derivatives_nav": _pick(nav_row, "nav phái sinh cuối kỳ"),
            "margin_debt": _pick(nav_row, "dư nợ cho vay cơ sở cuối kỳ") or _pick(daily_row, "dư nợ cho vay cuối kỳ"),
            "overdue_debt": _pick(daily_row, "dư nợ quá hạn cuối kỳ"),
            "aum": _pick(daily_row, "aum cuối kỳ"),
            "total_toi": _pick(toi_row, "tổng toi"),
            "positions": positions,
            "source_files": {kind: path.name for kind, path in files.items()},
            "privacy": {
                "customer_name_included": False,
                "account_number_included": False,
                "lookup_tokens_persisted": False,
            },
        }
        return snapshot

    def context_block(self, *, account_token: str, customer_token: str | None = None) -> str:
        snapshot = self.lookup(account_token=account_token, customer_token=customer_token)
        return (
            "## Private customer portfolio snapshot\n"
            "This de-identified snapshot was resolved locally from operator-trusted Excel exports. "
            "Treat it as the authoritative holdings/NAV input. Never request, infer, or print the "
            "customer name, account number, or lookup tokens.\n\n"
            "```json\n" + json.dumps(snapshot, ensure_ascii=False, indent=2) + "\n```"
        )
